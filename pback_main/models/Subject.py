from django.db import models

import os
import openpyxl as xlsx


class SubjectManager(models.Manager):
    def _get_subjects_start_column(self, worksheet):
        for i in range(1, 15):
            if worksheet['A' + str(i)].value == 'Дисциплина':
                for j in range(i + 1, 20):
                    if worksheet['A' + str(j)].value != None:
                        return j

    def _get_course_and_semester(self, worksheet):
        course = 0
        semester = 0
        for row in worksheet.iter_rows(min_col=1, min_row=1):
            for col in row:
                try:
                    if col.value == 'Курс':
                        course = col.column
                    if col.value == 'Семестр':
                        semester = col.column
                except:
                    continue
        return course, semester

    def set_objects_from_excel(self, path, owner):

        # delete old subjects
        self.filter(owner=owner).delete()

        wb = xlsx.load_workbook(path, data_only=True)
        ws = wb.active

        subjects_starts = self._get_subjects_start_column(ws)
        # TODO: counting semester in ZAO
        course, semester = self._get_course_and_semester(ws)

        # если ячейка строковая и без цвета - это предмет
        subjects = []
        for i in ws.iter_rows(min_row=subjects_starts, min_col=1, max_col=1):
            if i[0].data_type == 's' and not isinstance(i[0].font.color.rgb, str):
                current_subj = Subject()
                current_subj.row_number = i[0].row
                current_subj.name = i[0].value
                if semester != 0:
                    current_subj.semester = ws.cell(column=semester, row=i[0].row).value
                    if current_subj.semester is None:
                        current_subj.semester = 0
                else:
                    current_subj.semester = 0
                if course != 0:
                    current_subj.course = ws.cell(column=course, row=i[0].row).value
                    if current_subj.course is None:
                        current_subj.course = 0
                else:
                    current_subj.course = 0
                current_subj.owner = owner
                subjects.append(current_subj)
        self.bulk_create(subjects)

    def get_objects_json(self, owner):
        subjects = self.filter(owner=owner)
        json = []
        for subject in subjects:
            subj = {}
            subj['number'] = subject.row_number
            subj['subject'] = subject.name
            subj['semester'] = subject.semester
            subj['course'] = subject.course
            json.append(subj)
        return json


class Subject(models.Model):
    semester = models.CharField(max_length=3, blank=False)
    course = models.CharField(max_length=3, blank=False)
    name = models.CharField(max_length=100, blank=False)
    row_number = models.IntegerField(null=False)
    owner = models.ForeignKey('pback_auth.User', null=True, default=None, on_delete=models.SET_NULL)

    objects = SubjectManager()

    def __str__(self):
        return 'Subject {0} for {1} course {2} semester'.format(self.name, self.course, self.semester)
